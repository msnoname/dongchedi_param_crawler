# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html

# useful for handling different item types with a single interface
import sys
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import json
from tqdm import tqdm


MONTH = datetime.now().strftime('%Y%m')
BOLD = Font(bold=True)


class BrandPipeline(object):

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active

    def open_spider(self, spider):
        self.ws.append(['code_brand', 'brand', 'url'])
        self.ws.append(['品牌编码', '品牌', '网址'])
        for row in self.ws['1:2']:
            for cell in row:
                cell.font = BOLD

    def close_spider(self, spider):
        job = spider.name
        xlsx_path = f'data/{MONTH}/{job}/{job}_{MONTH}.xlsx'
        self.wb.save(xlsx_path)
        print('Success!')

    def process_item(self, item, spider):
        line = [item.get(key, None) for key in ['code_brand', 'brand', 'url']]
        self.ws.append(line)
        return item


class ParamPipeline(object):

    def __init__(self):
        self.start_time = None
        self.end_time = None

    def open_spider(self, spider):
        self.start_time = datetime.now()

    def close_spider(self, spider):
        print("--------------------------------------------------")
        print("Reading data from jsonlines...")
        job = sys.argv[-1].split('/')[-1]
        jsonl_path = f'data/{MONTH}/{job}/{job}_{MONTH}.jsonl'
        xlsx_path = f'data/{MONTH}/{job}/{job}_{MONTH}.xlsx'
        fail_path = f'data/{MONTH}/{job}/fail.json'
        wb = Workbook()
        ws = wb.active
        for i in range(len(spider.state["columns"])):
            ws.cell(row=1, column=i + 1, value=list(spider.state["columns"].keys())[i])
            ws.cell(row=2, column=i + 1, value=list(spider.state["columns"].values())[i])
        for row in ws['1:2']:
            for cell in row:
                cell.font = BOLD
        succeed = 0
        duplicated = 0
        with open(jsonl_path, 'r') as f:
            car_type_set = set()
            for line in tqdm(f):
                item = json.loads(line)
                if item["code_car_type"] not in car_type_set:
                    car_type_set.add(item["code_car_type"])
                    xlsx_line = []
                    for key in spider.state["columns"].keys():
                        value = item.get(key, None)
                        if \
                                key in {'market_time'} \
                                or not value \
                                or isinstance(value, (int, float)) \
                                or (len(value) > 1 and value[0] == '0' and value[1] != "."):
                            pass
                        elif value == '未知':
                            value = None
                        else:
                            try:
                                value = int(value)
                            except ValueError:
                                try:
                                    value = float(value)
                                except ValueError:
                                    pass
                        xlsx_line += [value]
                    ws.append(xlsx_line)
                    succeed += 1
                else:
                    duplicated += 1
        num_cols = len(spider.state["columns"])
        # print('Number of Columns:', num_cols)
        ref_str = "A2:{}2".format(get_column_letter(num_cols))
        ws.auto_filter.ref = ref_str  # 开启筛选功能
        print("Exporting to excel...")
        wb.save(xlsx_path)
        print("Excel saved successfully!")
        print("Spider closed: %s" % spider.name)
        print("--------------------------------------------------")
        self.end_time = datetime.now()
        time_diff = self.end_time - self.start_time
        spider.state['total_time_cost'] = spider.state.get('total_time_cost', timedelta()) + time_diff
        print(f"Time Cost: {time_diff} (Total: {spider.state['total_time_cost']})")
        # crawled = spider.state.get('crawled', 0)
        # print(f"Success: {succeed} = {crawled}(crawled) - {duplicated}(duplicated)")
        print(f"Success: {succeed}")
        with open(fail_path, 'w') as file:
            json.dump(spider.state.get('fail'), file)
        fail_car_series = spider.state.get('fail', {}).get('car_series', {})
        fail_car_type = spider.state.get('fail', {}).get('car_type', {})
        print(f"Failure: {len(fail_car_series)} car series, {len(fail_car_type)} car types")
        print("Finished!")

    def process_item(self, item, spider):
        item["dealer_price"] = float(item["dealer_price"].replace("万", "")) \
            if item.get("dealer_price") and item["dealer_price"] != "暂无报价" else None
        item["official_price"] = float(item["official_price"].replace("万", "")) \
            if item.get("official_price") and item["official_price"] != "暂无报价" else None
        item["market_time_year"] = int(item['market_time'].split('.')[0]) if item.get('market_time') else None
        item["market_time_month"] = int(item['market_time'].split('.')[1]) if item.get('market_time') else None
        return item


class DongchediPipeline:
    def process_item(self, item, spider):
        return item
